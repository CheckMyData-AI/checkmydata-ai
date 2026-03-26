import asyncio
import io
import json
import logging
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db
from app.core.audit import audit_log
from app.core.rate_limit import limiter
from app.services.batch_service import BatchService
from app.services.connection_service import ConnectionService
from app.services.membership_service import MembershipService

logger = logging.getLogger(__name__)

router = APIRouter()
_svc = BatchService()
_conn_svc = ConnectionService()
_membership_svc = MembershipService()


class BatchQueryItem(BaseModel):
    sql: str = Field(max_length=50000)
    title: str = Field(max_length=200)


class BatchExecuteRequest(BaseModel):
    project_id: str
    connection_id: str
    title: str = Field(max_length=200)
    queries: list[BatchQueryItem] = Field(default_factory=list)
    note_ids: list[str] | None = Field(None, max_length=100)


class BatchResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: str
    user_id: str
    project_id: str
    connection_id: str
    title: str
    queries_json: str
    note_ids_json: str | None
    status: str
    results_json: str | None
    created_at: datetime | None
    completed_at: datetime | None


@router.post("/execute", response_model=dict, status_code=202)
@limiter.limit("10/minute")
async def execute_batch(
    request: Request,
    body: BatchExecuteRequest,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await _membership_svc.require_role(db, body.project_id, user["user_id"], "viewer")

    conn = await _conn_svc.get(db, body.connection_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")
    if conn.project_id != body.project_id:
        raise HTTPException(status_code=400, detail="Connection does not belong to this project")

    if not body.queries and not body.note_ids:
        raise HTTPException(status_code=400, detail="Provide at least one query or note_id")

    queries_dicts = [q.model_dump() for q in body.queries]
    batch = await _svc.create_batch(
        db,
        user_id=user["user_id"],
        project_id=body.project_id,
        connection_id=body.connection_id,
        title=body.title,
        queries=queries_dicts,
        note_ids=body.note_ids,
    )

    audit_log(
        "batch.create",
        user_id=user["user_id"],
        project_id=body.project_id,
        resource_type="batch",
        resource_id=batch.id,
    )

    def _on_batch_done(t: asyncio.Task[None]) -> None:
        if t.cancelled():
            return
        exc = t.exception()
        if exc:
            logger.error(
                "Batch %s failed: %s",
                batch.id,
                exc,
                exc_info=(type(exc), exc, exc.__traceback__),
            )

    task = asyncio.create_task(
        _svc.execute_batch(batch.id, body.connection_id, user_id=user["user_id"])
    )
    task.add_done_callback(_on_batch_done)

    return {"batch_id": batch.id, "status": "pending"}


@router.get("/{batch_id}", response_model=BatchResponse)
async def get_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    batch = await _svc.get_batch(db, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if batch.user_id != user["user_id"]:
        raise HTTPException(status_code=403, detail="Not your batch")
    await _membership_svc.require_role(db, batch.project_id, user["user_id"], "viewer")
    return batch


@router.get("", response_model=list[BatchResponse])
async def list_batches(
    project_id: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await _membership_svc.require_role(db, project_id, user["user_id"], "viewer")
    return await _svc.list_batches(db, project_id, user["user_id"])


@router.delete("/{batch_id}")
@limiter.limit("20/minute")
async def delete_batch(
    request: Request,
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    batch = await _svc.get_batch(db, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if batch.user_id != user["user_id"]:
        raise HTTPException(status_code=403, detail="Not your batch")

    await _svc.delete_batch(db, batch_id)
    audit_log(
        "batch.delete",
        user_id=user["user_id"],
        project_id=batch.project_id,
        resource_type="batch",
        resource_id=batch_id,
    )
    return {"ok": True}


def _safe_sheet_name(title: str, idx: int) -> str:
    name = title[:28] if len(title) > 28 else title
    invalid = ["\\", "/", "*", "?", ":", "[", "]"]
    for ch in invalid:
        name = name.replace(ch, "_")
    return f"{idx + 1}_{name}" if name else f"Query_{idx + 1}"


@router.post("/{batch_id}/export")
@limiter.limit("10/minute")
async def export_batch(
    request: Request,
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    batch = await _svc.get_batch(db, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if batch.user_id != user["user_id"]:
        raise HTTPException(status_code=403, detail="Not your batch")
    if not batch.results_json:
        raise HTTPException(status_code=400, detail="No results to export")

    results = json.loads(batch.results_json)
    wb = Workbook()
    wb.remove(wb.active)

    for idx, res in enumerate(results):
        sheet_name = _safe_sheet_name(res.get("title", f"Query {idx + 1}"), idx)
        ws = wb.create_sheet(title=sheet_name[:31])

        if res.get("status") == "failed":
            ws.append(["Error", res.get("error", "Unknown error")])
            continue

        columns = res.get("columns", [])
        rows = res.get("rows", [])
        if columns:
            ws.append(columns)
        for row in rows:
            ws.append(row)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Empty")
        ws.append(["No results"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = batch.title.replace('"', "'")[:50]
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="batch_{safe_title}.xlsx"'},
    )
